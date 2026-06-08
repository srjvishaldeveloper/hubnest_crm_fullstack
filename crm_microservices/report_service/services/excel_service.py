import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_excel_report(data: dict, title: str):
    wb = openpyxl.Workbook()
    
    # Sheet 1: KPIs
    ws1 = wb.active
    ws1.title = "KPI Summary"
    
    # Colors & Fonts
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    alt_fill = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # Title Row
    ws1["A1"] = f"Job Nest CRM: {title} Executive Summary"
    ws1["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    ws1.row_dimensions[1].height = 25
    
    # Table Headers
    ws1["A3"] = "Metric"
    ws1["B3"] = "Value"
    ws1["A3"].fill = header_fill
    ws1["B3"].fill = header_fill
    ws1["A3"].font = header_font
    ws1["B3"].font = header_font
    ws1["A3"].alignment = Alignment(horizontal="left")
    ws1["B3"].alignment = Alignment(horizontal="right")
    
    kpis = data.get("kpis", {})
    row_num = 4
    for k, v in kpis.items():
        ws1.cell(row=row_num, column=1, value=k.replace("_", " ").title()).font = regular_font
        
        val_cell = ws1.cell(row=row_num, column=2, value=v)
        val_cell.font = regular_font
        if isinstance(v, float):
            val_cell.number_format = "$#,##0.00" if "rate" not in k.lower() else "0.0%"
            
        if row_num % 2 == 1:
            ws1.cell(row=row_num, column=1).fill = alt_fill
            val_cell.fill = alt_fill
            
        ws1.cell(row=row_num, column=1).border = thin_border
        val_cell.border = thin_border
        row_num += 1
        
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Sheet 2: Raw data table
    table_rows = data.get("table_data", [])
    if table_rows:
        ws2 = wb.create_sheet(title="Detailed Data")
        
        # Headers
        cols = list(table_rows[0].keys())
        for c_idx, col_name in enumerate(cols, 1):
            cell = ws2.cell(row=1, column=c_idx, value=col_name.replace("_", " ").title())
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
            
        # Data
        for r_idx, row_dict in enumerate(table_rows, 2):
            for c_idx, col_name in enumerate(cols, 1):
                val = row_dict[col_name]
                cell = ws2.cell(row=r_idx, column=c_idx, value=val)
                cell.font = regular_font
                cell.border = thin_border
                
                if isinstance(val, float):
                    cell.number_format = "$#,##0.00"
                
                if r_idx % 2 == 1:
                    cell.fill = alt_fill
                    
        for col in ws2.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws2.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
